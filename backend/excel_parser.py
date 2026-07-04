import openpyxl
from io import BytesIO
from typing import List, Dict, Tuple, Any

def get_excel_sheets(file_content: bytes) -> List[str]:
    """
    Returns the list of sheet names in the provided excel file bytes.
    """
    try:
        wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True)
        return wb.sheetnames
    except Exception as e:
        raise ValueError(f"Impossibile leggere il file Excel: {str(e)}")

def parse_warehouse_excel(
    file_content: bytes, 
    sheet_name: str,
    col_sku: str = "Sku",
    col_qty: str = "Qta Tot.",
    col_desc: str = "Descrizione Sku",
    col_lotto: str = "Lotto"
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """
    Parses the warehouse stock file and returns:
    - a list of valid stock rows (dict containing: sku, description, lotto, qty_total)
    - a list of anomalies discovered during parsing
    """
    valid_rows = []
    anomalies = []
    
    try:
        wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Foglio '{sheet_name}' non trovato nel file Excel.")
        
        sheet = wb[sheet_name]
        
        # Read header row (first row)
        header = [str(cell.value).strip() if cell.value is not None else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        
        # Identify columns with fallback
        # SKU
        try:
            sku_idx = next(i for i, h in enumerate(header) if h.lower() == col_sku.lower())
        except StopIteration:
            try:
                sku_idx = next(i for i, h in enumerate(header) if h.lower() == 'sku')
            except StopIteration:
                raise ValueError(f"Colonna SKU '{col_sku}' non trovata nella riga di intestazione (colonne presenti: {', '.join(filter(None, header))}).")
            
        # Qty
        try:
            qty_idx = next(i for i, h in enumerate(header) if h.lower() == col_qty.lower())
        except StopIteration:
            try:
                qty_idx = next(i for i, h in enumerate(header) if h.lower() in ('qta tot.', 'qty tot', 'quantità', 'qta', 'quantity'))
            except StopIteration:
                raise ValueError(f"Colonna Quantità '{col_qty}' non trovata nella riga di intestazione (colonne presenti: {', '.join(filter(None, header))}).")
            
        # Optional columns
        desc_idx = next((i for i, h in enumerate(header) if h.lower() == col_desc.lower()), None)
        if desc_idx is None:
            desc_idx = next((i for i, h in enumerate(header) if h.lower() in ('descrizione sku', 'descrizione', 'description')), None)
            
        lotto_idx = next((i for i, h in enumerate(header) if h.lower() == col_lotto.lower()), None)
        if lotto_idx is None:
            lotto_idx = next((i for i, h in enumerate(header) if h.lower() in ('lotto', 'lot')), None)
        
        valid_rows = []
        seen_skus = set()
        
        # Iterate data rows starting from row 2
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # If the entire row is empty, save it as a spacer row to replicate original layout
            if not any(cell is not None for cell in row):
                spacer_sku = f"__spacer_{row_num}"
                valid_rows.append({
                    "sku": spacer_sku,
                    "description": "",
                    "lotto": "",
                    "qty_total": 0.0
                })
                continue
                
            sku_val = row[sku_idx] if sku_idx < len(row) else None
            qty_val = row[qty_idx] if qty_idx < len(row) else None
            
            # Clean SKU
            if sku_val is None or str(sku_val).strip() == "":
                anomalies.append({
                    "source": "stock_import",
                    "record_key": f"Riga {row_num}",
                    "anomaly_type": "missing_sku",
                    "message": f"Riga {row_num}: SKU vuota o mancante, riga ignorata."
                })
                continue
                
            sku = str(sku_val).strip()
            
            # Parse quantity
            try:
                if qty_val is None:
                    raise ValueError("Quantità mancante")
                qty_total = float(qty_val)
                if qty_total < 0:
                    anomalies.append({
                        "source": "stock_import",
                        "record_key": sku,
                        "anomaly_type": "negative_quantity",
                        "message": f"La SKU '{sku}' ha una quantità negativa alla riga {row_num}: {qty_total}."
                    })
            except Exception as e:
                anomalies.append({
                    "source": "stock_import",
                    "record_key": sku,
                    "anomaly_type": "invalid_quantity",
                    "message": f"Riga {row_num} (SKU: {sku}): Quantità non numerica o mancante ('{qty_val}'). Riga ignorata."
                })
                continue
                
            description = str(row[desc_idx]).strip() if desc_idx is not None and desc_idx < len(row) and row[desc_idx] is not None else None
            lotto = str(row[lotto_idx]).strip() if lotto_idx is not None and lotto_idx < len(row) and row[lotto_idx] is not None else None
            
            # Check for duplicate SKUs for anomaly reporting
            if sku in seen_skus:
                anomalies.append({
                    "source": "stock_import",
                    "record_key": sku,
                    "anomaly_type": "duplicate_sku",
                    "message": f"La SKU '{sku}' compare più volte nel file di stock (riga {row_num}). Questa riga viene importata separatamente."
                })
            else:
                seen_skus.add(sku)
                
            valid_rows.append({
                "sku": sku,
                "description": description or "",
                "lotto": lotto or "",
                "qty_total": qty_total
            })
            
        # Clean up trailing empty/spacer rows if they exceed 5
        last_real_idx = -1
        for i in range(len(valid_rows) - 1, -1, -1):
            if not valid_rows[i]["sku"].startswith("__spacer_"):
                last_real_idx = i
                break
                
        if last_real_idx != -1:
            num_trailing_spacers = len(valid_rows) - 1 - last_real_idx
            if num_trailing_spacers > 5:
                valid_rows = valid_rows[:last_real_idx + 1]
        else:
            if len(valid_rows) > 5:
                valid_rows = []
            
        return valid_rows, anomalies
        
    except Exception as e:
        if not isinstance(e, ValueError):
            raise ValueError(f"Errore durante il parsing del file Excel di stock: {str(e)}")
        raise e

def parse_associations_excel(file_content: bytes) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """
    Parses the association file and returns:
    - a list of associations (product_id, sku, qty_required)
    - a list of anomalies discovered during parsing
    
    Rule: No header row, reading from row 1.
    Column A: product_id (integer)
    Column B: comma-separated SKU list.
    """
    associations = []
    anomalies = []
    
    try:
        wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
        sheet = wb.active # Use active sheet (first sheet)
        
        for row_num, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if not any(cell is not None for cell in row):
                continue
                
            prod_id_val = row[0] if len(row) > 0 else None
            sku_list_val = row[1] if len(row) > 1 else None
            
            # Parse product ID
            if prod_id_val is None or str(prod_id_val).strip() == "":
                anomalies.append({
                    "source": "associations_import",
                    "record_key": f"Riga {row_num}",
                    "anomaly_type": "missing_product_id",
                    "message": f"Riga {row_num}: Product ID mancante, riga ignorata."
                })
                continue
                
            try:
                product_id = int(float(str(prod_id_val).strip()))
            except Exception:
                anomalies.append({
                    "source": "associations_import",
                    "record_key": str(prod_id_val),
                    "anomaly_type": "invalid_product_id",
                    "message": f"Riga {row_num}: Product ID non valido ({prod_id_val}), riga ignorata."
                })
                continue
                
            # Parse SKU list
            if sku_list_val is None or str(sku_list_val).strip() == "":
                anomalies.append({
                    "source": "associations_import",
                    "record_key": str(product_id),
                    "anomaly_type": "empty_sku_list",
                    "message": f"Il prodotto '{product_id}' (Riga {row_num}) ha un elenco SKU vuoto."
                })
                continue
                
            sku_str = str(sku_list_val)
            # Split by comma and clean
            skus_split = [s.strip() for s in sku_str.split(',') if s.strip() != '']
            
            if not skus_split:
                anomalies.append({
                    "source": "associations_import",
                    "record_key": str(product_id),
                    "anomaly_type": "empty_sku_list",
                    "message": f"Il prodotto '{product_id}' ha un elenco SKU vuoto dopo la pulizia."
                })
                continue
                
            # Count occurrences of each SKU
            sku_counts = {}
            for sku in skus_split:
                sku_counts[sku] = sku_counts.get(sku, 0) + 1
                
            # Generate output
            for sku, qty in sku_counts.items():
                associations.append({
                    "product_id": product_id,
                    "sku": sku,
                    "qty_required": qty
                })
                
        # Consolidate duplicate (product_id, sku) pairs across different rows
        consolidated_assocs = {}
        for assoc in associations:
            key = (assoc["product_id"], assoc["sku"])
            if key in consolidated_assocs:
                prev_qty = consolidated_assocs[key]["qty_required"]
                consolidated_assocs[key]["qty_required"] += assoc["qty_required"]
                anomalies.append({
                    "source": "associations_import",
                    "record_key": f"{assoc['product_id']}",
                    "anomaly_type": "duplicate_association",
                    "message": f"Il prodotto composto '{assoc['product_id']}' ha associazioni duplicate per la SKU '{assoc['sku']}'. Le quantità richieste sono state consolidate (Sommate: {prev_qty} + {assoc['qty_required']} = {consolidated_assocs[key]['qty_required']})."
                })
            else:
                consolidated_assocs[key] = assoc
                
        return list(consolidated_assocs.values()), anomalies
        
    except Exception as e:
        raise ValueError(f"Errore durante il parsing del file Excel di associazioni: {str(e)}")
