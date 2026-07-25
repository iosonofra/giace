from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from backend.models import ImportBatch, ProductComponent


def export_associations_excel(db) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Associazioni"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:B1"
    sheet.append([
        "Product ID",
        "Componenti SKU (ripetute per quantità)",
    ])

    active_batch = (
        db.query(ImportBatch)
        .filter(
            ImportBatch.file_type == "associations",
            ImportBatch.is_active.is_(True),
        )
        .first()
    )
    if active_batch:
        components = (
            db.query(ProductComponent)
            .filter(
                ProductComponent.import_batch_id == active_batch.id,
            )
            .order_by(
                ProductComponent.product_id.asc(),
                ProductComponent.id.asc(),
            )
            .all()
        )
        grouped = {}
        for component in components:
            grouped.setdefault(component.product_id, []).extend(
                [component.sku] * component.qty_required
            )
        for product_id, skus in grouped.items():
            sheet.append([product_id, ", ".join(skus)])

    header_fill = PatternFill("solid", fgColor="6366F1")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 24
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 64

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
